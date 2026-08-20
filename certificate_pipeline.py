"""
Core certificate-generation pipeline.

generate_certificates(xlsx_path, date_str=None) -> pdf_path

Every layout number and fix here was arrived at through trial-and-error in
an earlier interactive session (documented in README.md, "Known pitfalls").
Do not "simplify" the draw order or the rectangle sizes without re-reading
that section -- several of them look redundant but exist to fix specific
visual bugs (clipped descenders, signature tails overlapping text, one
white cover-rect erasing another element that was already drawn).
"""
import io
import os
import re
import logging
from datetime import datetime

import openpyxl
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase.pdfmetrics import stringWidth
from PIL import Image

import config
from name_render import render_name_image

log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# 1. Read + clean the roster out of the xlsx
# ---------------------------------------------------------------------------
def parse_score(value):
    """Best-effort parse of a 'x / y' quiz score string; returns -1 if not
    parseable (e.g. Excel mis-auto-formatted it as a date)."""
    if isinstance(value, str) and "/" in value:
        try:
            return float(value.split("/")[0].strip())
        except ValueError:
            return -1
    return -1


def find_name_column(header_row):
    for i, cell in enumerate(header_row):
        if cell and any(k in str(cell) for k in ("ชื่อ", "Name", "name")):
            return i
    return None


def find_factory_column(header_row):
    for i, cell in enumerate(header_row):
        if cell and any(k in str(cell) for k in ("โรงงาน", "Factory", "Company")):
            return i
    return None


def extract_date_string(wb, ws, display_format="%d %B, %Y"):
    """Try, in order:
      1. A sheet name that looks like a date, e.g. '15Jun26'.
      2. A cell in row 1 literally containing a parseable date.
    Returns (date_str, was_found). If nothing was found, date_str falls
    back to today's date but was_found=False, so the caller can flag this
    loudly (an automated agent must never silently guess a wrong date on a
    real certificate)."""
    m = re.match(r"(\d{1,2})([A-Za-z]{3})(\d{2,4})", ws.title)
    if m:
        day, mon, yr = m.groups()
        yr = ("20" + yr) if len(yr) == 2 else yr
        try:
            dt = datetime.strptime(f"{day} {mon} {yr}", "%d %b %Y")
            return dt.strftime(display_format), True
        except ValueError:
            pass

    for row in ws.iter_rows(max_row=1, values_only=True):
        for cell in row:
            if isinstance(cell, datetime):
                return cell.strftime(display_format), True

    log.warning("Could not find a date in the xlsx; falling back to today's date.")
    return datetime.today().strftime(display_format), False


def load_roster(xlsx_path):
    """Returns (names: list[str], date_str: str, date_was_found: bool)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header, data_rows = rows[0], rows[1:]

    name_col = find_name_column(header)
    factory_col = find_factory_column(header)
    if name_col is None:
        # Fall back to the layout seen in the sample files: name is
        # column C (index 3) when there's no recognizable header.
        name_col = 3

    order, best_score = [], {}
    for row in data_rows:
        if name_col >= len(row):
            continue
        raw_name = row[name_col]
        if not raw_name:
            continue
        name = str(raw_name).replace("\u200b", "").strip()
        if not name:
            continue

        if factory_col is not None and factory_col < len(row):
            factory = row[factory_col]
            if factory is not None and str(factory).strip().lower() == "เฉลย":
                continue  # answer-key row, not a real participant

        name = config.apply_honorific(name)
        score = parse_score(row[2]) if len(row) > 2 else -1
        if name not in best_score:
            order.append(name)
            best_score[name] = score
        else:
            best_score[name] = max(best_score[name], score)

    date_str, date_was_found = extract_date_string(wb, ws)
    return order, date_str, date_was_found


# ---------------------------------------------------------------------------
# 2. Signature placement helper
# ---------------------------------------------------------------------------
def fitted_box_top_anchored(img_path, box, max_height):
    """Fit image (preserving aspect) anchored to the TOP of box, capped at
    max_height, so a long signature tail doesn't extend down into text
    below it."""
    bx, by, bw, bh = box
    top_y = by + bh
    im = Image.open(img_path)
    iw, ih = im.size
    scale = min(bw / iw, max_height / ih)
    w, h = iw * scale, ih * scale
    x = bx + (bw - w) / 2
    y = top_y - h
    return x, y, w, h


# ---------------------------------------------------------------------------
# 3. Render one certificate page (as a reportlab overlay merged onto a
#    fresh copy of the template page)
# ---------------------------------------------------------------------------
def render_certificate_page(template_bytes, name, date_str, signatories, work_dir):
    reader = PdfReader(io.BytesIO(template_bytes))
    page = reader.pages[0]

    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(config.PAGE_W, config.PAGE_H))

    # --- Phase 1: white-out every region being replaced, ALL BEFORE any
    # new content is drawn. (A rect drawn after new content risks silently
    # erasing part of that content if their areas overlap -- this bit us
    # once with the date's leading digit.) ---
    c.setFillColorRGB(1, 1, 1)
    c.rect(*config.NAME_COVER_RECT, fill=1, stroke=0)
    c.rect(*config.DATE_COVER_RECT, fill=1, stroke=0)
    c.rect(*config.LEFT_COVER_RECT, fill=1, stroke=0)
    c.rect(*config.RIGHT_COVER_RECT, fill=1, stroke=0)

    # --- Phase 2: draw all new content on top ---
    c.setFillColorRGB(0, 0, 0)

    # Name: rendered via RAQM (see name_render.py) so Thai combining
    # vowels/tone marks stack correctly -- reportlab's native drawString
    # cannot do this.
    img_path, w_name, h_name, top_to_baseline = render_name_image(
        name, config.NAME_FONT_PATH, config.NAME_FONT_SIZE, work_dir
    )
    img_x = config.CENTER_X - w_name / 2
    img_y = config.NAME_BASELINE_Y - (h_name - top_to_baseline)
    c.drawImage(img_path, img_x, img_y, w_name, h_name, mask="auto")

    # Date
    c.setFont(config.DATE_FONT, config.DATE_FONT_SIZE)
    w_date = stringWidth(date_str, config.DATE_FONT, config.DATE_FONT_SIZE)
    c.drawString(config.CENTER_X - w_date / 2, config.DATE_BASELINE_Y, date_str)

    # Left + right signatory blocks
    for side, box, x0 in (
        ("left", config.LEFT_SIG_BOX, config.LEFT_X0),
        ("right", config.RIGHT_SIG_BOX, config.RIGHT_X0),
    ):
        sig = signatories[side]
        for text, size, baseline in sig["lines"]:
            c.setFont(config.FOOTER_FONT, size)
            c.drawString(x0, baseline, text)
        sx, sy, sw, sh = fitted_box_top_anchored(sig["signature_image"], box, sig["max_height"])
        c.drawImage(sig["signature_image"], sx, sy, sw, sh, mask="auto")

    c.save()
    buf.seek(0)
    overlay_page = PdfReader(buf).pages[0]
    page.merge_page(overlay_page)
    return page


# ---------------------------------------------------------------------------
# 4. Top-level entry point
# ---------------------------------------------------------------------------
def generate_certificates(xlsx_path, out_dir, date_str_override=None, signatories=None):
    """Builds one PDF containing one certificate page per unique roster
    name. Returns the output PDF path."""
    os.makedirs(out_dir, exist_ok=True)
    work_dir = os.path.join(out_dir, "name_imgs")

    signatories = signatories or config.DEFAULT_SIGNATORIES

    names, parsed_date, date_was_found = load_roster(xlsx_path)
    if not names:
        raise ValueError("No names found in the xlsx -- check the file format / column headers.")
    date_str = date_str_override or parsed_date

    with open(config.TEMPLATE_PATH, "rb") as f:
        template_bytes = f.read()

    writer = PdfWriter()
    for name in names:
        page = render_certificate_page(template_bytes, name, date_str, signatories, work_dir)
        writer.add_page(page)

    out_path = os.path.join(out_dir, "Certificates.pdf")
    with open(out_path, "wb") as f:
        writer.write(f)

    log.info("Generated %d certificates -> %s", len(names), out_path)
    date_flag_note = None if (date_str_override or date_was_found) else (
        "WARNING: no date was found in the xlsx (no dated sheet name, no date cell in row 1) "
        f"-- defaulted to today's date ({date_str}). Verify this is correct before distributing."
    )
    return out_path, names, date_str, date_flag_note
